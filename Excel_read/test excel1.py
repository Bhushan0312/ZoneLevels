import openpyxl as opxl
from tkinter import *
print('Starting application.......')
xl=opxl.load_workbook("ZONE LEVELS.xlsx")
sh = xl["Sheet1"]
global val
def exception_lb():
    try:
        clear_TB()
    except:
        pass
    exception_lb.l1 = Label(root, text='Please enter valid VWAP value', bg='#293659', font=("arial", 7, "bold"), fg='yellow')
    exception_lb.l1.place(x=298, y=284)
def clear_TB():
    T.delete(0,'end')
    try:
        exception_lb.l1.destroy()
    except:
        pass
    try:
        compare.l2.destroy()
    except:
        pass
    try:
        compare.l3.destroy()
    except:
        pass
    try:
        compare.l4.destroy()
    except:
        pass
    try:
        compare.l5.destroy()
    except:
        pass
    try:
        compare.l6.destroy()
    except:
        pass
    try:
        compare.l7.destroy()
    except:
        pass
    try:
        compare.l8.destroy()
    except:
        pass
    try:
        compare.l9.destroy()
    except:
        pass
def stop():
    root.destroy()
    print('Closing Application.....')
def compare(val):
    try:
        clear_TB()
    except:
        pass
    less = []
    greater = []
    for k in range(1, (sh.max_column) + 1):
        for l in range(1, (sh.max_row) + 1):
            if (val > round(float(sh.cell(l, k).value),3)):
                less.append(round(float(sh.cell(l, k).value),3))
            elif (val < round(float(sh.cell(l, k).value),3)):
                greater.append(round(float(sh.cell(l, k).value),3))

    try:
        less.sort()
        less = list(dict.fromkeys(less))
        compare.l2 = Label(root, text=less[-1], bg='#d40000', font=("arial", 15, "bold"), fg='White')
        compare.l2.place(x=750, y=376)#Sell
        try:
            compare.l3 = Label(root, text=less[-2], bg='#a40000', font=("arial", 15, "bold"), fg='white')
            compare.l3.place(x=750, y=455)#sell target 1
            compare.l4 = Label(root, text=less[-3], bg='#840000', font=("arial", 15, "bold"), fg='white')
            compare.l4.place(x=750, y=530)  # sell target 2
        except:
            compare.l5 = Label(root, text='SELL value is smallest value : ' + less[-1], bg='#151A30',font=("arial", 10, "bold"), fg='white')
            compare.l5.place(x=679, y=407)  # SELL value is smallest value
    except:
        pass
    try:
        greater.sort()
        greater = list(dict.fromkeys(greater))
        compare.l6 = Label(root, text=str(round(greater[0],3)), bg='#0daf01', font=("arial", 15, "bold"), fg='White')
        compare.l6.place(x=260, y=375)#buy value
        try:
            compare.l7 = Label(root, text=greater[1], bg='#0a8101', font=("arial", 15, "bold"), fg='White')
            compare.l7.place(x=260, y=452)#buy target 1
            compare.l8 = Label(root, text=greater[2], bg='#086200', font=("arial", 15, "bold"), fg='White')
            compare.l8.place(x=260, y=529)  # buy target 2
        except:
            compare.l9 = Label(root, text='BUY value is greatest value : '+greater[0], bg='#151A30', font=("arial", 10, "bold"), fg='white')
            compare.l9.place(x=46, y=407)#BUY value is greatest value
    except:
        pass

def search(val):

    try:
        val=float(T.get())
        findless = []
        findgreat = []
        for i in range(1,(sh.max_column)+1):
            for j in range(1,(sh.max_row)+1):
                if(sh.cell(j,i).value == None):
                    sh.cell(j,i).value="0"
                    xl.save("ZONE LEVELS.xlsx")
        for i in range(1, (sh.max_column) + 1):
             for j in range(1, (sh.max_row) + 1):
                if(val>= float(sh.cell(j,i).value)):
                        findless.append(round(float(sh.cell(j,i).value),3))
                elif(val<= float(sh.cell(j,i).value)):
                        findgreat.append(round(float(sh.cell(j,i).value),3))
        try:
            findless.sort()
            findless=list(dict.fromkeys(findless))
            findgreat.sort()
            findgreat=list(dict.fromkeys(findgreat))
            vall=val-findless[-1]
            try:
                    valg = findgreat[0] - val
            except:
                    valg = 0
            if(vall==0.0000000 and valg==0.000000 ):
                    compare(val)
            elif(vall <= valg):
                    compare(findless[-1])
            elif (vall >= valg):
                    compare(findgreat[0])
        except:
          exception_lb()
          pass

    except ValueError:
        exception_lb()

def bt():
    try:
        search(float(T.get()))
    except ValueError:
        exception_lb()

root = Tk()
root.title("ZONE LEVELS")
root.geometry("980x695+50+50")
root.configure(bg="#293659")
root.resizable(0,0)
main1 = PhotoImage(file='icon/mainp.png')
img = Label(root,image=main1,borderwidth=0)
img.place(x=0,y=0,)
T = Entry(root,width=21,font=("Arial",25),bg='#f0f0f0',borderwidth=0)
T.place(x=288,y=234)
T.bind('<Return>',search)
loadimg = PhotoImage(file='icon/search.png')
Search_bt = Button(root,image=loadimg,bd=0,bg='#293659',activebackground='#293659',command=bt)
Search_bt.place(x=722, y=232)
clrimg = PhotoImage(file='icon/clear.png')
clear_bt=Button(root,image=clrimg,bd=0,bg='#293659',activebackground='#293659',command=clear_TB)
clear_bt.place(x=853,y=231)
close_img=PhotoImage(file='icon/Close.png')
close_bt = Button(root,image=close_img,bd=0,bg='#151A30',activebackground='#151A30',command=stop)
close_bt.place(x=774,y=636)
root.mainloop()
