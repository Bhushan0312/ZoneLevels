import openpyxl as opxl
xl=opxl.load_workbook("ZONE LEVELS.xlsx")
sh = xl["Sheet1"]
less=[]
greater=[]
findless=[]
findgreat=[]
print(sh.max_row)
val= float(input("CMP ="))
def compare(val):
    for k in range(1, (sh.max_column) + 1):
        for l in range(1, (sh.max_row) + 1):
            if (int(val) > int(sh.cell(l, k).value)):
                less.append(int(sh.cell(l, k).value))
            elif (int(val) < int(sh.cell(l, k).value)):
                greater.append(int(sh.cell(l, k).value))
    return less
    return greater

for i in range(1,(sh.max_column)+1):
    for j in range(1,(sh.max_row)+1):
        if(sh.cell(j,i).value == None):
            sh.cell(j,i).value="0"
            xl.save("ZONE LEVELS.xlsx")
for i in range(1, (sh.max_column) + 1):
     for j in range(1, (sh.max_row) + 1):
        if(val>= int(sh.cell(j,i).value) or val >= float(sh.cell(j,i).value)):
                findless.append(int(sh.cell(j,i).value))
        elif(val<= int(sh.cell(j,i).value) or val <= float(sh.cell(j,i).value)):
                findgreat.append(int(sh.cell(j,i).value))
findless.sort()
findless=list(dict.fromkeys(findless))
findgreat.sort()
findgreat=list(dict.fromkeys(findgreat))
vall=val-findless[-1]
try:
    valg=findgreat[0]-val
except:
    valg=0
if(vall==0 and valg==0 ):
    compare(val)
elif(vall>=valg):
    compare(findgreat[0])
elif(vall<=valg):
    compare(findless[-1])

try:
    less.sort()
    less=list(dict.fromkeys(less))
    print("SELL",less[-1])
    print("SELL TARGET ",less[-2])
except:
    print("there is no previous value of",(val))
try:
    greater.sort()
    less = list(dict.fromkeys(less))
    print("BUY",greater[0])
    try:
         print("TARGET FOR BUY",greater[1])
    except:
        print("BUY value is greatest value :",greater[0])
except:
    print("there is no next value of",(val))
